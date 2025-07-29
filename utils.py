import email
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime

import imaplib
from imaplib import IMAP4_SSL
import json
import os

import jdatetime
import openpyxl
import pandas as pd

from mailbox import Message

from typing import Any, Dict, Generator, List, Tuple


def connect_to_email(mail_addr: str, mail_passwd: str) -> IMAP4_SSL:
    """
    Connect to an email account using IMAP over SSL.

    Parameters:
        mail_addr (str): Email address (e.g., user@gmail.com)
        mail_passwd (str): App-specific password for the email account

    Returns:
        IMAP4_SSL: An authenticated IMAP connection object
    """
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(mail_addr, mail_passwd)
    return mail

def fetch_email_addresses(emails_file_path: str) -> List[str]:
    """
    Reads a CSV file and extracts a list of email addresses.

    Parameters:
        emails_file_path (str): Path to the CSV file containing email addresses.

    Returns:
        List[str]: A list of email addresses from the 'email_address' column.

    Raises:
        ValueError: If the 'email_address' column is missing.
    """
    email_addresses_df = pd.read_csv(emails_file_path)
    if 'email_address' not in email_addresses_df.columns:
        raise ValueError("CSV file must contain an 'email_address' column.")
    return list(email_addresses_df['email_address'])

def get_filtered_emails(mail: IMAP4_SSL, mails_file_path: str) -> Generator[Tuple[str, Message], None, None]:
    """
    Generator that yields full email message objects (email.message.Message)
    for emails from allowed addresses.

    Parameters:
        mail (IMAP4_SSL): Authenticated IMAP connection
        mails_file_path (str): Path to emails file
    Yields:
        Message: Full parsed email message object
    """
    mail.select("inbox")
    status, messages = mail.search(None, 'ALL')
    if status != 'OK':
        return
    mail_ids = messages[0].split()
    emails = fetch_email_addresses(mails_file_path)
    for mail_id in reversed(mail_ids):
        status, mail_content = mail.fetch(mail_id, "(RFC822)")
        if status != 'OK':
            continue
        for response_part in mail_content:
            if isinstance(response_part, tuple):
                mail_body = email.message_from_bytes(response_part[1])
                from_ = mail_body.get("From", "")
                if "<" in from_ and ">" in from_:
                    sender_email = from_.split("<")[1].split(">")[0]
                else:
                    sender_email = from_
                if sender_email.strip() in emails:
                    yield mail_id, mail_body


def extract_mail_info(mail: IMAP4_SSL, mails_file_path: str) -> Generator[Dict[str, Any], None, None]:
    """
    Extracts metadata from filtered emails that are not replies or forwards.

    Parameters:
        mail (IMAP4_SSL): Authenticated IMAP connection to an email inbox
        mails_file_path (str): Path to the CSV file containing whitelisted sender email addresses

    Yields:
        dict: A dictionary containing email metadata such as sender, subject, date, time, etc.
    """
    mails = get_filtered_emails(mail, mails_file_path)
    for mail_id, mail_body in mails:
        if mail_body.get('In-Reply-To') or mail_body.get('References'):
            continue
        py_content = get_all_mail_py_attached_content(mail_body)
        py_files_num = count_py_attachments(mail_body)
        name, email_addr = parseaddr(mail_body.get('From', ''))
        try:
            email_date = parsedate_to_datetime(mail_body.get('Date', '')).date()
            email_time = parsedate_to_datetime(mail_body.get('Date', '')).time()
        except Exception as e:
            print(e)
            email_date, email_time = "", ""
        subject_raw = mail_body.get("Subject", "")
        subject, encoding = decode_header(subject_raw)[0]
        if isinstance(subject, bytes):
            subject = subject.decode(encoding or "utf-8", errors="ignore")
        mail_info = {
            'mail_id': str(mail_id),
            'mail_sender': name,
            'mail_address': email_addr,
            'mail_time': email_time.strftime('%H:%M:%S'),
            'mail_date':  jdatetime.date.fromgregorian(date=email_date).isoformat(),
            'mail_subject': subject,
            'py_content': py_content,
            'py_files': py_files_num
        }
        yield mail_info


def add_mail_info_to_excel(email_info: Dict[str, Any]) -> None:
    """
    Adds a new email record to an Excel file, avoiding duplicates by mail_id.

    Parameters:
        email_info (Dict): Dictionary containing email metadata
    """
    mails_info_file_path = os.getenv("MAILS_INFO_FILE_PATH")
    if not mails_info_file_path:
        raise EnvironmentError("MAILS_INFO_FILE_PATH not found in environment variables.")
    if not os.path.exists(mails_info_file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = list(email_info.keys())
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(mails_info_file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if 'mail_id' in headers:
            mail_id_column = headers.index('mail_id')
            existing_ids = [row[mail_id_column] for row in ws.iter_rows(min_row=2, values_only=True)]
            if email_info['mail_id'] in existing_ids:
                wb.close()
                return
        else:
            raise ValueError("'mail_id' column not found in Excel file.")
    row_data = [email_info.get(header, "") for header in headers]
    ws.append(row_data)
    wb.save(mails_info_file_path)
    wb.close()


def count_py_attachments(mail_body: email.message.Message) -> int:
    """
    Counts how many attached files with `.py` extension an email has.

    Parameters:
        mail_body (Message): Parsed email message

    Returns:
        int: Number of `.py` attachments
    """
    count = 0
    for part in mail_body.walk():
        content_disposition = part.get("Content-Disposition", "")
        if "attachment" in content_disposition.lower():
            filename = part.get_filename()
            if filename and filename.lower().endswith('.py'):
                count += 1
    return count

def get_all_mail_py_attached_content(mail_body: email.message.Message) -> str:
    """
    Extracts and concatenates the contents of all `.py` attachments in an email.

    Parameters:
        mail_body (Message): Parsed email message

    Returns:
        str: Combined string of all `.py` file contents
    """
    contents = []

    for part in mail_body.walk():
        content_disposition = part.get("Content-Disposition", "")
        if "attachment" in content_disposition.lower():
            filename = part.get_filename()
            if filename and filename.lower().endswith('.py'):
                file_data = part.get_payload(decode=True)
                try:
                    content = file_data.decode('utf-8')
                except UnicodeDecodeError:
                    content = file_data.decode('latin1')  # fallback
                contents.append(f"# === File: {filename} ===\n{content.strip()}")

    return "\n\n".join(contents)


def load_config(config_path: str = "config.json") -> dict:
    """
    Load config file from a JSON path.

    Parameters:
        config_path (str): Path of config file.
    Returns:
        dict: Parsed JSON config file content.
    Raises:
        FileNotFoundError: If the config file does not exist.
        json.JSONDecodeError: If the file content is not valid JSON.
    """
    with open(config_path, "r", encoding="utf-8") as config_file:
        return json.load(config_file)


def fetch_question() -> Tuple[Any, Any]:
    """
    Reads a CSV file specified by the environment variable 'QUESTION_FILE_PATH',
    and returns the first row's values from the 'question' and 'session' columns.

    Raises:
        ValueError: If the environment variable 'QUESTION_FILE_PATH' is not set.
        KeyError: If 'question' or 'session' columns are missing in the CSV file.
        pd.errors.EmptyDataError: If the CSV file is empty.

    Returns:
        Tuple[Any, Any]: A tuple containing the first 'question' and 'session' values.
    """
    question_file_path = os.getenv('QUESTION_FILE_PATH')
    if question_file_path is None:
        raise ValueError("Environment variable 'QUESTION_FILE_PATH' is not set.")

    questions = pd.read_csv(question_file_path)

    if 'question' not in questions.columns or 'session' not in questions.columns:
        raise KeyError("Columns 'question' or 'session' not found in the CSV file.")

    question = questions['question'].iloc[0]
    session = questions['session'].iloc[0]

    return question, session
